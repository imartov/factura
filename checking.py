import os
from openpyxl import load_workbook


def checking_files():
    path = f"{os.getcwd()}\\xlsx"
    files_list = os.listdir(path)
    # print(f"\nВсего найдено файлов: {len(files_list)}")

    none_rows = ''
    validation_files = []
    validation_sheets = []
    for file_name in files_list:
        try:
            validation_files.append(file_name)
            wb = load_workbook(filename=f"{path}\\{file_name}")
            print(f'\nПроверка файла: {file_name}')
            print(f"Всего найдено листов: {len(wb.sheetnames) - 1}")

            for sheet_name in wb.sheetnames:
                if sheet_name == 'values_for_lists':
                    continue
                else:
                    validation_sheets.append(sheet_name)
                    sheet_active = wb[sheet_name]
                    print(f'Проверка листа: {sheet_name}')

                    i = 0
                    for row in sheet_active.iter_rows(min_row=12, max_col=3):
                        products_name_sheet = row[0]
                        count_sheet = row[1]
                        price_sheet = row[2]

                        if not products_name_sheet.value and not count_sheet.value and not price_sheet.value:
                            break
                        elif products_name_sheet.value and count_sheet.value and price_sheet.value:
                            pass
                        else:
                            row = str(row).replace('(', '').replace(')', '\n')
                            print(f'Файл: {file_name}, Лист: {sheet_name}, недопустимые значения строки {row}')
                            none_rows += f'Файл: {file_name}, Лист: {sheet_name}, строка: {row}'

                        i += 1

        except Exception as ex:
            print(f'Исключение: {ex}\n')
            continue

    if none_rows:
        with open(f'{os.getcwd()}\\none_cells.txt', 'w', encoding='utf-8') as file:
            file.write(none_rows)
        print(f'Найдены пустые значения в следующих строках:\n{none_rows}')
        print('Заполните поля и введите "next" для повторной проверки')
        if input_cell() == 'next':
            try:
                os.remove(f"{os.getcwd()}\\none_cells.txt")
            except Exception:
                pass
        checking_files()
    else:
        print('\nПроверка документа пройдена успешно, пустые поля отсутствуют.')
        return validation_files


def input_cell():
    menu = ['next']
    input_data = input('Место для ввода: ')
    if input_data not in menu:
        print(f'\nВведены недопустимые данные\nВозможные варианты для ввода:')
        print(', '.join(menu))
        input_cell()
    return input_data


def waiting_for_checking():
    menu = ['stop', 'next']
    print('\nВыполнение программы приостановлено\nПроверьте введенные данные\n\nДля возобновление введите "next"\nДля остановки введите "stop"')
    input_data = input('Место для ввода: ')
    if input_data not in menu:
        print('\nВведено недопустимое значение')
        waiting_for_checking()
    return input_data


def main():
    checking_files()
    input_cell()
    waiting_for_checking()


if __name__ == "__main__":
    main()
