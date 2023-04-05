import selenium
from selenium import webdriver
from selenium.common import ElementClickInterceptedException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import Select
import os
from openpyxl import load_workbook
from fake_useragent import UserAgent
from login_facturowo import input_login, get_saved_login_data, get_temp_login_data
from tqdm import tqdm, trange
from googletrans import Translator
from checking import *


def get_factura():
    validation_files = checking_files()
    if "saved_login.json" in os.listdir(os.getcwd()):
        pass
    else:
        input_login()

    # user_agent = UserAgent()

    # options = webdriver.ChromeOptions()
    # options.add_argument(f"user-agent={user_agent.random}")

    # driver = webdriver.Chrome(options=options)
    driver = webdriver.Chrome()
    driver.maximize_window()
    url = "https://www.fakturowo.pl/logowanie"
    driver.get(url)

    sleep(3)

    # login
    login_mail = driver.find_element(By.XPATH, f"//input[@name='email']")
    login_mail.send_keys(Keys.CONTROL, "a")
    login_mail.send_keys(Keys.DELETE)

    login_password = driver.find_element(By.XPATH, f"//input[@name='haslo']")
    login_password.send_keys(Keys.CONTROL, "a")
    login_password.send_keys(Keys.DELETE)

    if "saved_login.json" in os.listdir(os.getcwd()):
        email, password = get_saved_login_data()
        login_mail.send_keys(email)
        login_password.send_keys(password)

    else:
        email, password = get_temp_login_data()
        login_mail.send_keys(email)
        login_password.send_keys(password)

    # button_login = driver.find_element(By.XPATH, f"//button[@name='login']")
    # button_login.click()
    login_password.send_keys(Keys.RETURN)

    sleep(2)

    # if entered login data are wrong
    err = False
    while not err:
        try:
            driver.find_element(By.XPATH, "//div[@class='alert alert-danger']")
            print("\nВведенные логин и пароль не верны. Пожалуйста, повторите ввод пароля")
            input_login()
            sleep(1)

            login_mail = driver.find_element(By.XPATH, f"//input[@name='email']")
            login_mail.send_keys(Keys.CONTROL, "a")
            login_mail.send_keys(Keys.DELETE)

            login_password = driver.find_element(By.XPATH, f"//input[@name='haslo']")
            login_password.send_keys(Keys.CONTROL, "a")
            login_password.send_keys(Keys.DELETE)

            if "saved_login.json" in os.listdir(os.getcwd()):
                email, password = get_saved_login_data()
                login_mail.send_keys(email)
                login_password.send_keys(password)
            else:
                email, password = get_temp_login_data()
                login_mail.send_keys(email)
                login_password.send_keys(password)

            button_login = driver.find_element(By.XPATH, f"//button[@name='login']")
            button_login.click()
        except Exception:
            err = True

    # delete temp login and password
    try:
        os.remove(f"{os.getcwd()}\\temp_login.json")
    except Exception:
        pass

    # add button for creating new document
    new_document_button = driver.find_element(By.XPATH, f"//a[@class='btn btn-xs btn-primary']")
    new_document_button.click()

    sleep(1.5)

    prev_sheet_count_products = 0

    # getting input excel files
    path = f"{os.getcwd()}\\xlsx"
    for file in validation_files:
        print(f'\nИзвлечение данных из файла: {file}')
        # getting sheets in every excel file
        wb = load_workbook(filename=f"{path}\\{file}")
        print(f"Всего найдено листов: {len(wb.sheetnames) - 1}.")


        for sheet_name in wb.sheetnames:

            # if prev CMR had error and the fields didn't refresh
            if prev_sheet_count_products:
                try:
                    for pev_count_product in range(prev_sheet_count_products - 1):
                        delete_product_button = driver.find_element(By.XPATH, f"//a[@onclick='deleteRow({pev_count_product + 1},0,arr);']")
                        delete_product_button.click()
                        sleep(0.5)
                except Exception:
                    pass

            if sheet_name == 'values_for_lists':
                continue
            else:
                print(f'\nИзвлечение данных из листа: {sheet_name}')

                # select active sheet
                sheet_active = wb[sheet_name]
                # select data from sheet from excel file
                document_type_sheet = str(sheet_active['B1'].value)
                buyer_name_sheet = str(sheet_active['B2'].value)
                buyer_nip_sheet = str(sheet_active['B3'].value)
                buyer_address_sheet = str(sheet_active['B4'].value)
                buyer_place_sheet = str(sheet_active['B5'].value)
                tax_sheet = str(sheet_active['B6'].value)
                currency_sheet = str(sheet_active['B7'].value)
                measure_sheet = str(sheet_active['B8'].value)

                # filling field 'Dokument'
                document_type_select = Select(driver.find_element(By.ID, 'rodzaj'))
                document_type_select.select_by_visible_text(document_type_sheet)

                # filling fields about buyer
                buyer_name_field = driver.find_element(By.NAME, "nabywca[nazwa]")
                buyer_name_field.send_keys(Keys.CONTROL, "a")
                buyer_name_field.send_keys(Keys.DELETE)
                buyer_name_field.send_keys(buyer_name_sheet)

                sleep(0.5)

                buyer_nip_field = driver.find_element(By.NAME, "nabywca[nip]")
                buyer_nip_field.send_keys(Keys.CONTROL, "a")
                buyer_nip_field.send_keys(Keys.DELETE)
                buyer_nip_field.send_keys(buyer_nip_sheet)

                sleep(0.5)

                buyer_address_field = driver.find_element(By.ID, 'ulica_nabywca')
                buyer_address_field.send_keys(Keys.CONTROL, "a")
                buyer_address_field.send_keys(Keys.DELETE)
                buyer_address_field.send_keys(buyer_address_sheet)

                sleep(0.5)

                buyer_place_field = driver.find_element(By.ID, 'miasto_nabywca')
                buyer_place_field.send_keys(Keys.CONTROL, "a")
                buyer_place_field.send_keys(Keys.DELETE)
                buyer_place_field.send_keys(buyer_place_sheet)

                sleep(0.5)

                # filling field 'Waluta'
                currency_field = Select(driver.find_element(By.ID, 'waluta'))
                currency_field.select_by_visible_text(currency_sheet)

                sleep(1)

                # for row in tqdm(range(1, sheet_active.max_column, 1), desc=sheet_name, unit="product", dynamic_ncols=True):
                i = 0
                for row in sheet_active.iter_rows(min_row=12, max_col=3, values_only=True):

                    if row[0]:

                        # select data about products
                        product_name_sheet_ru = str(row[0])
                        count_sheet = str(row[1])
                        price_sheet = str(row[2])

                        translator = Translator()
                        product_name_sheet_pl = translator.translate(text=product_name_sheet_ru, src='ru', dest='pl')

                        # filling field about product
                        product_name_site = driver.find_element(By.ID, f"nazwa_{i}")
                        product_name_site.send_keys(Keys.CONTROL, "a")
                        product_name_site.send_keys(Keys.DELETE)
                        product_name_site.send_keys(product_name_sheet_pl.text.capitalize())

                        sleep(0.5)

                        # select option of measure
                        select_measure = Select(driver.find_element(By.ID, f"jm_{i}"))
                        select_measure.select_by_visible_text(measure_sheet)

                        sleep(0.5)

                        # select field of count
                        count_site = driver.find_element(By.ID, f"ilosc_{i}")
                        count_site.send_keys(Keys.CONTROL, "a")
                        count_site.send_keys(Keys.DELETE)
                        count_site.send_keys(count_sheet)

                        sleep(0.5)

                        # select field of price
                        price_netto_site = driver.find_element(By.XPATH, f"//input[@id='cena_netto_{i}']")
                        price_netto_site.send_keys(Keys.CONTROL, "a")
                        price_netto_site.send_keys(Keys.DELETE)
                        price_netto_site.send_keys(price_sheet)

                        sleep(0.5)

                        # filling field 'Stawka VAT'
                        try:
                            tax_field = Select(driver.find_element(By.ID, f'stawka_vat_{i}'))
                            if document_type_sheet == 'Eksport towarów (poza UE)':
                                tax_value = '0'
                                tax_field.select_by_value(tax_value)
                            else:
                                tax_field.select_by_value(tax_sheet)
                        except Exception:
                            print(f'\nЗначение для поля "Stawka VAT" не соответствует выбранному типу документа для поля "Dokument"\n')

                        sleep(0.5)

                        add_product_button = driver.find_element(By.XPATH, "//a[@onclick='addRow(arr)']")
                        add_product_button.click()

                        sleep(0.5)
                        i += 1
                    # remove excess product's field
                    else:
                        delete_product_button = driver.find_element(By.XPATH, f"//a[@onclick='deleteRow({i},0,arr);']")
                        delete_product_button.click()
                        break

                try:
                    # again remove excess product's field
                    delete_product_button = driver.find_element(By.XPATH, f"//a[@onclick='deleteRow({i},0,arr);']")
                    delete_product_button.click()

                    # just click for filling field 'Wartość netto'
                    # just_click = driver.find_element(By.XPATH, "//input[@name='wartosc_netto[0]']")
                    # just_click.click()
                except Exception:
                    pass

                if waiting_for_checking() == 'stop':
                    driver.close()
                    driver.quit()
                    return print('Принудительное завершение программы')

                save_factura_button = driver.find_element(By.XPATH, "//button[@id='pobierz_i_zapisz']")
                driver.execute_script("arguments[0].click();", save_factura_button)

                # if error message is displayed
                error_message = driver.find_element(By.XPATH, '/html/body/div[1]/main/section[2]/div/div/div[1]/div')
                if error_message.is_displayed():
                    print('\nНекоторые поля на странице не корректны либо пусты')

                    def save_document():
                        if waiting_for_checking() == 'stop':
                            driver.close()
                            driver.quit()
                            return print('Принудительное завершение программы')

                        refresh_currency_page = Select(driver.find_element(By.XPATH, '//*[@id="waluta"]'))
                        refresh_currency_page.select_by_visible_text('PLN')
                        sleep(1)
                        refresh_currency_page = Select(driver.find_element(By.XPATH, '//*[@id="waluta"]'))
                        refresh_currency_page.select_by_visible_text(currency_sheet)

                        save_factura_button = driver.find_element(By.XPATH, "//button[@id='pobierz_i_zapisz']")
                        driver.execute_script("arguments[0].click();", save_factura_button)

                        sleep(0.5)

                        error_message = driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/main/section[2]/div/div/div[1]/div')
                        if error_message.is_displayed():
                            print('\nНекоторые поля на странице не заполнены либо пусты')
                            save_document()

                    save_document()

                refresh_page = driver.find_element(By.XPATH, "//a[@class='link-gray']")
                refresh_page.click()

                sleep(1)

    driver.close()
    driver.quit()
    return print("\nProgram execution completed successfully")


def main():
    get_factura()


if __name__ == "__main__":
    main()
