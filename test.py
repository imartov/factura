from openpyxl import load_workbook
import os


def get_data_xlsx():
    path = "D:\python_projects\\factura\\xlsx"
    files_list = os.listdir(path)

    for index, file in enumerate(files_list, start=1):
        wb = load_workbook(filename=f"{path}\{file}")
        sheet_one = wb.active

        for row in range(1, sheet_one.max_column, 1):
            product_name = sheet_one[1][row].value
            count = sheet_one[3][row].value
            price_netto = sheet_one[4][row].value

            print(f"{product_name}\n{count}\n{price_netto}\n")


def main():
    get_data_xlsx()


if __name__ == "__main__":
    main()
