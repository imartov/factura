import os
from openpyxl import load_workbook


def checking_files():
    path = f"{os.getcwd()}\\xlsx_test"
    files_list = os.listdir(path)
    print(f"\nВсего найдено файлов: {len(files_list)}.")

    for file in files_list:
        print(f"\nИзвлечен файл {file}.")

        # getting sheets in every excel file
        wb = load_workbook(filename=f"{path}\\{file}")
        print(f"Всего найдено листов: {len(wb.sheetnames)}.\n")

        for sheet in wb.sheetnames:
            sheet_active = wb[sheet]

            print(f"max_row: {sheet_active.max_row}, max_col: {sheet_active.max_column}\n")

            for row in sheet_active.iter_rows(max_row=sheet_active.max_row, max_col=sheet_active.max_column, min_row=2):
                products_name_sheet = row[0].value
                count_sheet = row[1].value
                price_sheet = row[2].value

                if products_name_sheet is None:
                    print(f"Имя товара в ячейке {row[0]} не заполнено.")
                if count_sheet is None:
                    print(f"Количество товара {products_name_sheet} в ячейке {row[1]} не заполнено.")
                if price_sheet is None:
                    print(f"Стоимость товара {products_name_sheet} в ячейке {row[2]} не заполнено.")


def input_cell():
    print("Введите '1', если хотите заполнить ячейку в программе")
    print("Введите '2', если хотите завершить программу и отредактировать файл самостоятельно")
    input_cell = input("")


def main():
    checking_files()
    input_cell()


if __name__ == "__main__":
    main()
